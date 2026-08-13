import base64
import html
import io
import re
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from service.cypress_runner import screenshot_absolute_path
_HEADER_FILL = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
_HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
_PASS_FILL = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
_FAIL_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
_PASS_FONT = Font(color='166534', bold=True)
_FAIL_FONT = Font(color='991B1B', bold=True)
_TOTAL_FONT = Font(bold=True)
_WRAP_TOP = Alignment(wrap_text=True, vertical='top')
_CENTER = Alignment(horizontal='center', vertical='center')
_CENTER_TOP = Alignment(horizontal='center', vertical='top')
_EXAMPLE_SUFFIX_RE = re.compile('\\s*\\(example\\s*#?\\d+\\)\\s*$', re.IGNORECASE)
_TC_DELIMITER_RE = re.compile('^\\s*#[─\\-\\s]*TC-\\d+\\s*:\\s*(.+?)\\s*[─\\-]*\\s*$')
_SCENARIO_LINE_RE = re.compile('^\\s*Scenario(?:\\s+Outline)?\\s*:\\s*(.+?)\\s*$')
_PLACEHOLDER_RE = re.compile('<[^<>]+>')

def _scenario_name(test_case_title: str) -> str:
    return _EXAMPLE_SUFFIX_RE.sub('', test_case_title or '').strip()

def _parse_tc_blocks(feature_text: str) -> list[dict]:
    blocks: list[dict] = []
    pending_title: str | None = None
    for line in (feature_text or '').splitlines():
        delim_match = _TC_DELIMITER_RE.match(line)
        if delim_match:
            pending_title = delim_match.group(1).strip()
            continue
        scen_match = _SCENARIO_LINE_RE.match(line)
        if scen_match:
            raw_name = scen_match.group(1).strip()
            pattern = re.escape(raw_name)
            for token in _PLACEHOLDER_RE.findall(raw_name):
                pattern = pattern.replace(re.escape(token), '(.+?)', 1)
            blocks.append({'title': pending_title, 'raw_name': raw_name, 'match_re': re.compile(f'^{pattern}$')})
            pending_title = None
            continue
        if line.strip() and (not line.strip().startswith('@')):
            pending_title = None
    return blocks

def _match_block(scenario_name: str, blocks: list[dict]) -> tuple[dict | None, str | None]:
    for block in blocks:
        m = block['match_re'].match(scenario_name)
        if m:
            condition = m.group(1).strip() if m.groups() else None
            return (block, condition)
    return (None, None)

def build_report_xlsx(module: str, run_results: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'QC Report'
    headers = ['Module', 'Screen', 'Scenario', 'Test Case', 'Pass/Fail', 'Passed', 'Failed', 'Total']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    ws.freeze_panes = 'A2'
    row = 2
    for run in run_results:
        screen = run.get('slug') or '(unnamed)'
        screen_start_row = row
        flat_scenarios: list[dict] = []
        for suite in run.get('suites') or []:
            flat_scenarios.extend(suite.get('scenarios') or [])
        screen_pass = 0
        screen_fail = 0
        if not flat_scenarios:
            screen_fail += 1
            ws.cell(row=row, column=1, value=module)
            ws.cell(row=row, column=2, value=screen)
            ws.cell(row=row, column=3, value='(no scenarios executed)')
            ws.cell(row=row, column=4, value='(no scenarios executed)')
            c_status = ws.cell(row=row, column=5, value='Failed')
            c_status.alignment = _CENTER
            c_status.fill = _FAIL_FILL
            c_status.font = _FAIL_FONT
            row += 1
        else:
            blocks = _parse_tc_blocks(run.get('feature_text') or '')
            i, n = (0, len(flat_scenarios))
            while i < n:
                name_i = _scenario_name(flat_scenarios[i].get('name') or '(untitled)')
                block_i, _ = _match_block(name_i, blocks)
                mergeable = block_i is not None and block_i['title'] is not None
                group: list[dict] = [flat_scenarios[i]]
                j = i + 1
                if mergeable:
                    while j < n:
                        name_j = _scenario_name(flat_scenarios[j].get('name') or '(untitled)')
                        block_j, _ = _match_block(name_j, blocks)
                        if block_j is block_i:
                            group.append(flat_scenarios[j])
                            j += 1
                        else:
                            break
                scenario_start_row = row
                for scen in group:
                    name = _scenario_name(scen.get('name') or '(untitled)')
                    block, condition = _match_block(name, blocks)
                    state = scen.get('state') or 'failed'
                    passed = state == 'passed'
                    if passed:
                        screen_pass += 1
                    else:
                        screen_fail += 1
                    if block and block['title']:
                        scenario_text = block['raw_name']
                        testcase_text = f'{block['title']} - {condition}' if condition else block['title']
                    else:
                        scenario_text = name
                        testcase_text = name
                    ws.cell(row=row, column=1, value=module)
                    ws.cell(row=row, column=2, value=screen)
                    c_scenario = ws.cell(row=row, column=3, value=scenario_text)
                    c_testcase = ws.cell(row=row, column=4, value=testcase_text)
                    c_status = ws.cell(row=row, column=5, value='Passed' if passed else 'Failed')
                    c_scenario.alignment = _WRAP_TOP
                    c_testcase.alignment = _WRAP_TOP
                    c_status.alignment = _CENTER
                    c_status.fill = _PASS_FILL if passed else _FAIL_FILL
                    c_status.font = _PASS_FONT if passed else _FAIL_FONT
                    row += 1
                scenario_end_row = row - 1
                if scenario_end_row > scenario_start_row:
                    ws.merge_cells(start_row=scenario_start_row, start_column=3, end_row=scenario_end_row, end_column=3)
                ws.cell(row=scenario_start_row, column=3).alignment = _WRAP_TOP
                i = j
        screen_end_row = row - 1
        merge_cols = (1, 2, 6, 7, 8)
        if screen_end_row > screen_start_row:
            for col in merge_cols:
                ws.merge_cells(start_row=screen_start_row, start_column=col, end_row=screen_end_row, end_column=col)
        for col in merge_cols:
            ws.cell(row=screen_start_row, column=col).alignment = _CENTER_TOP
        for col, value in ((6, screen_pass), (7, screen_fail), (8, screen_pass + screen_fail)):
            c = ws.cell(row=screen_start_row, column=col, value=value)
            c.font = _TOTAL_FONT
    widths = {'A': 16, 'B': 24, 'C': 42, 'D': 42, 'E': 16, 'F': 10, 'G': 10, 'H': 10}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
_BASE_CSS = '\n  * { box-sizing: border-box; margin: 0; padding: 0; }\n  body {\n    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;\n    color: #1f2937; background: #f9fafb; line-height: 1.55; padding: 32px;\n  }\n  .container { max-width: 960px; margin: 0 auto; }\n  header.report-header {\n    background: #fff; border: 1px solid #e5e7eb; border-radius: 12px;\n    padding: 24px 28px; margin-bottom: 20px;\n  }\n  header.report-header h1 { font-size: 22px; font-weight: 600; margin-bottom: 4px; }\n  header.report-header .meta { color: #6b7280; font-size: 13px; }\n  header.report-header .meta > span { margin-right: 16px; }\n\n  .shot-group {\n    background: #fff; border: 1px solid #e5e7eb; border-radius: 12px;\n    padding: 20px 24px; margin-bottom: 20px;\n  }\n  .shot-group > h2 { font-size: 15px; font-weight: 600; margin-bottom: 6px; }\n  .shot-group > .sub { font-size: 12px; color: #6b7280; margin-bottom: 14px; font-family: SFMono-Regular, Consolas, monospace; }\n  .shot { margin-top: 16px; padding-top: 16px; border-top: 1px solid #f3f4f6; }\n  .shot:first-child { border-top: none; padding-top: 0; margin-top: 0; }\n  .shot .caption { font-size: 13px; color: #374151; margin-bottom: 8px; font-weight: 500; }\n  .shot img { max-width: 100%; border: 1px solid #e5e7eb; border-radius: 6px; display: block; }\n\n  .err {\n    background: #fef2f2; border-left: 3px solid #fca5a5;\n    padding: 10px 14px; border-radius: 4px;\n    font-family: SFMono-Regular, Consolas, monospace; font-size: 12px;\n    color: #7f1d1d; white-space: pre-wrap; word-break: break-word;\n  }\n  footer { text-align: center; color: #9ca3af; font-size: 12px; margin-top: 30px; }\n  .empty {\n    background: #fff; border: 1px dashed #d1d5db; border-radius: 12px;\n    padding: 40px; text-align: center; color: #9ca3af; font-size: 14px;\n  }\n'

def _fmt_timestamp(iso: str | None) -> str:
    if not iso:
        return '—'
    try:
        dt = datetime.fromisoformat(iso.replace('Z', '+00:00'))
        return dt.astimezone(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')
    except (ValueError, TypeError):
        return iso

def _read_and_embed(rel_path: str) -> str | None:
    abs_path = screenshot_absolute_path(rel_path)
    try:
        data = abs_path.read_bytes()
    except OSError:
        return None
    b64 = base64.b64encode(data).decode('ascii')
    return f'data:image/png;base64,{b64}'

def _scenario_for_screenshot(run: dict, filename: str) -> str:
    stem = Path(filename).stem
    for suffix in [' (failed)', ' (attempt 2)', ' (attempt 3)']:
        if stem.endswith(suffix):
            stem = stem[:-len(suffix)]
    if ' -- ' in stem:
        return stem.split(' -- ', 1)[1]
    return stem

def build_screenshots_html(module: str, run_results: list[dict]) -> str:
    parts: list[str] = []
    parts.append("<!doctype html><html><head><meta charset='utf-8'>")
    parts.append(f'<title>Screenshots — {html.escape(module or 'run')}</title>')
    parts.append(f"<style>{_BASE_CSS}</style></head><body><div class='container'>")
    parts.append("<header class='report-header'>")
    parts.append(f'<h1>Screenshots — {html.escape(module or 'unknown module')}</h1>')
    total_shots = sum((len(r.get('screenshots') or []) for r in run_results))
    parts.append(f"<div class='meta'><span><strong>Total screenshots:</strong> {total_shots}</span>")
    parts.append(f'<span><strong>Across:</strong> {len(run_results)} screen(s)</span></div>')
    parts.append('</header>')
    if not total_shots:
        parts.append("<div class='empty'>No screenshots were captured during this run.<br/>Cypress captures screenshots automatically on scenario failure — a run with no failures produces none.</div>")
    else:
        for run in run_results:
            shots = run.get('screenshots') or []
            if not shots:
                continue
            slug = run.get('slug') or '(unnamed)'
            parts.append("<div class='shot-group'>")
            parts.append(f'<h2>{html.escape(slug)}</h2>')
            parts.append(f"<div class='sub'>{len(shots)} screenshot(s)</div>")
            for rel in shots:
                filename = Path(rel).name
                scenario_name = _scenario_for_screenshot(run, filename)
                data_url = _read_and_embed(rel)
                parts.append("<div class='shot'>")
                parts.append(f"<div class='caption'>{html.escape(scenario_name)}</div>")
                if data_url:
                    parts.append(f"<img src='{data_url}' alt='{html.escape(filename)}'/>")
                else:
                    parts.append(f"<div class='err'>Could not read screenshot file: {html.escape(rel)}</div>")
                parts.append('</div>')
            parts.append('</div>')
    parts.append(f'<footer>Generated by QC Test Console · {_fmt_timestamp(datetime.utcnow().isoformat() + 'Z')}</footer>')
    parts.append('</div></body></html>')
    return ''.join(parts)